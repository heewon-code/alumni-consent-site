#!/usr/bin/env python3
"""동문 개인정보 수집 웹 애플리케이션 (클라우드 배포 버전)"""

import os, json, base64, io, datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import psycopg2
import psycopg2.extras
import tornado.ioloop
import tornado.web
import tornado.escape
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ──────────────────────────────────────────
# 설정
# ──────────────────────────────────────────
pdfmetrics.registerFont(UnicodeCIDFont('HYSMyeongJo-Medium'))
FONT = 'HYSMyeongJo-Medium'
# bold/italic 변형 매핑 (CIDFont는 단일 weight이므로 모두 동일 폰트로 매핑)
from reportlab.lib.fonts import addMapping
addMapping('HYSMyeongJo-Medium', 0, 0, 'HYSMyeongJo-Medium')
addMapping('HYSMyeongJo-Medium', 1, 0, 'HYSMyeongJo-Medium')
addMapping('HYSMyeongJo-Medium', 0, 1, 'HYSMyeongJo-Medium')
addMapping('HYSMyeongJo-Medium', 1, 1, 'HYSMyeongJo-Medium')

# Render.com은 postgres:// 로 시작하는 URL을 제공하지만 psycopg2는 postgresql:// 필요
_db_url = os.environ.get('DATABASE_URL', '')
DATABASE_URL = _db_url.replace('postgres://', 'postgresql://', 1) if _db_url.startswith('postgres://') else _db_url
# 관리자 비밀번호 (환경변수 ADMIN_PASSWORD 로 변경 가능)
ADMIN_PASS      = os.environ.get('ADMIN_PASSWORD', 'duksung2026')
# 이메일 알림 설정
GMAIL_USER      = os.environ.get('GMAIL_USER', '')
GMAIL_APP_PASS  = os.environ.get('GMAIL_APP_PASSWORD', '')
NOTIFY_EMAIL    = os.environ.get('NOTIFY_EMAIL', 'hwkim@marusys.com')


# ──────────────────────────────────────────
# PostgreSQL 초기화
# ──────────────────────────────────────────
def get_conn():
    return psycopg2.connect(DATABASE_URL)


def init_db():
    con = get_conn()
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS submissions (
            id           SERIAL PRIMARY KEY,
            year         TEXT,
            name         TEXT,
            email        TEXT,
            phone        TEXT,
            address      TEXT,
            consent      TEXT,
            submitted_at TEXT,
            pdf_data     BYTEA
        )
    """)
    con.commit()
    cur.close()
    con.close()


def save_submission(data: dict, pdf_bytes: bytes = None):
    con = get_conn()
    cur = con.cursor()
    cur.execute(
        "INSERT INTO submissions (year,name,email,phone,address,consent,submitted_at,pdf_data)"
        " VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
        (data.get('year',''), data.get('name',''), data.get('email',''),
         data.get('phone',''), data.get('address',''),
         '동의' if data.get('consent')=='yes' else '미동의',
         datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
         psycopg2.Binary(pdf_bytes) if pdf_bytes else None)
    )
    con.commit()
    cur.close()
    con.close()


def send_notification_email(data: dict, pdf_bytes: bytes):
    """동문 제출 시 관리자 이메일로 PDF 알림 발송"""
    if not GMAIL_USER or not GMAIL_APP_PASS:
        return  # 환경변수 미설정 시 조용히 스킵
    try:
        name    = data.get('name', '동문')
        year    = data.get('year', '')
        consent = '동의' if data.get('consent') == 'yes' else '미동의'

        msg = MIMEMultipart()
        msg['From']    = GMAIL_USER
        msg['To']      = NOTIFY_EMAIL
        msg['Subject'] = f'[동문 개인정보 수집] {name} ({year}학번) 제출 완료'

        body = (
            f"동문이 개인정보 수집 동의서를 제출하였습니다.\n\n"
            f"  이름     : {name}\n"
            f"  학번     : {year}\n"
            f"  이메일   : {data.get('email','')}\n"
            f"  연락처   : {data.get('phone','')}\n"
            f"  주소     : {data.get('address','')}\n"
            f"  동의여부 : {consent}\n\n"
            f"서명이 포함된 PDF가 첨부되어 있습니다."
        )
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        if pdf_bytes:
            pdf_part = MIMEApplication(pdf_bytes, _subtype='pdf')
            pdf_part.add_header(
                'Content-Disposition', 'attachment',
                filename=f'개인정보동의서_{name}.pdf'
            )
            msg.attach(pdf_part)

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(GMAIL_USER, GMAIL_APP_PASS)
            smtp.sendmail(GMAIL_USER, NOTIFY_EMAIL, msg.as_string())
    except Exception as e:
        print(f'[이메일 오류] {e}')


def get_all_submissions():
    con = get_conn()
    cur = con.cursor()
    cur.execute(
        "SELECT id,year,name,email,phone,address,consent,submitted_at,"
        " (pdf_data IS NOT NULL) AS has_pdf FROM submissions ORDER BY id"
    )
    rows = cur.fetchall()
    cur.close()
    con.close()
    return rows


# ──────────────────────────────────────────
# PDF 생성
# ──────────────────────────────────────────
def make_consent_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        topMargin=12*mm, bottomMargin=20*mm,
        leftMargin=30*mm, rightMargin=30*mm)

    normal  = ParagraphStyle('ks_normal', fontName=FONT, fontSize=10, leading=18, wordWrap='CJK')
    title_s = ParagraphStyle('ks_title',  fontName=FONT, fontSize=16, leading=24, alignment=TA_CENTER, spaceAfter=4*mm)
    bold_s  = ParagraphStyle('ks_bold',   fontName=FONT, fontSize=10, leading=18, wordWrap='CJK', spaceAfter=2*mm)
    center_s= ParagraphStyle('ks_center', fontName=FONT, fontSize=10, leading=18, alignment=TA_CENTER, wordWrap='CJK')
    right_s = ParagraphStyle('ks_right',  fontName=FONT, fontSize=10, leading=18, alignment=TA_RIGHT, wordWrap='CJK')

    story = []
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph('개인정보 수집 및 이용 동의서', title_s))
    story.append(HRFlowable(width='100%', thickness=1.5, color=colors.black, spaceAfter=6*mm))

    story.append(Paragraph('1. 본인은 다음과 같이 본인의 개인정보를 수집&#8729;이용하는 것에 대하여 동의합니다.', bold_s))
    for line in [
        '가. 개인정보의 수집&#8729;이용자 : 덕성여자대학교',
        '나. 개인정보의 수집&#8729;이용 목적 : 동문 관리, 학교 행사 및 소식 안내 등',
        ('다. 수집&#8729;이용하려는 개인정보의 항목(필수정보)<br/>'
         '&#160;&#160;&#160;&#160;: 전공(학과), 입학연도(학번), 성명, 연락처(이메일, 핸드폰번호, 주소)'),
        '라. 개인정보의 보유 및 이용 기간 : 제공받은 자가 기록 보존이 필요한 시기까지 보유',
    ]:
        story.append(Paragraph(f'&#160;&#160;&#160;&#160;{line}', normal))
    story.append(Spacer(1, 4*mm))

    story.append(Paragraph('2. 개인정보 수집&#8729;이용을 거부할 권리 및 그에 따른 불이익', bold_s))
    story.append(Paragraph(
        '&#160;&#160;&#160;&#160;귀하는 위와 같은 개인정보 수집&#8729;이용에 동의하지 않을 수 있습니다.<br/>'
        '&#160;&#160;&#160;&#160;그러나 동의를 거부할 경우 동문 관리 및 학교 행사 및 소식 안내 등의 대상에서 제외될 수 있습니다.',
        normal))
    story.append(Spacer(1, 6*mm))

    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.grey, spaceAfter=4*mm))
    story.append(Paragraph('위와 같이 개인정보를 수집&#8729;이용하는 데 동의하십니까?', center_s))
    story.append(Spacer(1, 3*mm))

    agreed = data.get('consent') == 'yes'
    agree_t = Table([[
        Paragraph('■ 동의합니다' if agreed else '□ 동의합니다', normal),
        Paragraph('□ 동의하지 않습니다' if agreed else '■ 동의하지 않습니다', normal)
    ]], colWidths=[70*mm, 80*mm])
    agree_t.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER')]))
    story.append(agree_t)
    story.append(Spacer(1, 8*mm))

    today = datetime.date.today()
    story.append(Paragraph(f'{today.year} 년&#160;&#160; {today.month} 월&#160;&#160; {today.day} 일', center_s))
    story.append(Spacer(1, 8*mm))

    gray = colors.Color(0.93,0.93,0.93)
    info_data = [
        [Paragraph('전공(학과)', normal), Paragraph('유아교육과', normal),
         Paragraph('입학연도(학번)', normal), Paragraph(str(data.get('year','')), normal)],
        [Paragraph('성&#160;&#160;&#160;명', normal), Paragraph(data.get('name',''), normal),
         Paragraph('이&#160;&#160;메&#160;&#160;일', normal), Paragraph(data.get('email',''), normal)],
        [Paragraph('핸드폰', normal), Paragraph(data.get('phone',''), normal),
         Paragraph('', normal), Paragraph('', normal)],
        [Paragraph('주&#160;&#160;&#160;소', normal), Paragraph(data.get('address',''), normal),
         Paragraph('', normal), Paragraph('', normal)],
    ]
    info_t = Table(info_data, colWidths=[28*mm, 52*mm, 34*mm, 36*mm])
    info_t.setStyle(TableStyle([
        ('FONTNAME',(0,0),(-1,-1),FONT), ('FONTSIZE',(0,0),(-1,-1),10),
        ('GRID',(0,0),(-1,-1),0.5,colors.black),
        ('BACKGROUND',(0,0),(0,-1),gray), ('BACKGROUND',(2,0),(2,-1),gray),
        # SPAN 영역(핸드폰·주소 값 셀) 안의 회색 배경을 흰색으로 덮어 균일하게
        ('BACKGROUND',(1,2),(3,2),colors.white), ('BACKGROUND',(1,3),(3,3),colors.white),
        ('ALIGN',(0,0),(-1,-1),'CENTER'), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('TOPPADDING',(0,0),(-1,-1),5), ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('SPAN',(1,2),(3,2)), ('SPAN',(1,3),(3,3)),
    ]))
    story.append(info_t)
    story.append(Spacer(1, 6*mm))

    sig_b64 = data.get('signature','')
    if sig_b64 and sig_b64.startswith('data:image'):
        try:
            _, encoded = sig_b64.split(',', 1)
            from reportlab.platypus import Image as RLImage
            sig_io = io.BytesIO(base64.b64decode(encoded))
            sig_t = Table([[Paragraph('서&#160;&#160;&#160;명', normal), RLImage(sig_io, width=45*mm, height=18*mm)]],
                          colWidths=[28*mm, 50*mm])
            sig_t.setStyle(TableStyle([
                ('GRID',(0,0),(-1,-1),0.5,colors.black),
                ('BACKGROUND',(0,0),(0,0),gray),
                ('ALIGN',(0,0),(-1,-1),'CENTER'), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ]))
            story.append(sig_t)
            story.append(Spacer(1, 6*mm))
        except Exception:
            pass

    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.grey, spaceAfter=4*mm))
    story.append(Paragraph('덕성여자대학교 총장 귀하', right_s))
    doc.build(story)
    return buf.getvalue()


# ──────────────────────────────────────────
# Excel 내보내기
# ──────────────────────────────────────────
def export_excel() -> bytes:
    rows = get_all_submissions()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '동문명단'

    headers = ['순번','동문회 직위','입학연도(학번)','성명','이메일','핸드폰','주소','동의여부','제출일시']
    hfill = PatternFill('solid', fgColor='1A1A2E')
    hfont = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
    thin  = Side(style='thin', color='BBBBBB')
    bdr   = Border(left=thin,right=thin,top=thin,bottom=thin)

    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.fill=hfill; c.font=hfont
        c.alignment=Alignment(horizontal='center',vertical='center')
        c.border=bdr
    ws.row_dimensions[1].height = 22

    dfont = Font(name='맑은 고딕', size=10)
    for ri, row in enumerate(rows, 2):
        db_id, year, name, email, phone, address, consent, submitted_at, has_pdf = row
        vals = [ri-1,'동문',year,name,email,phone,address,consent,submitted_at]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(ri, ci, val)
            c.font=dfont
            c.alignment=Alignment(horizontal='center' if ci in (1,2,8,9) else 'left',vertical='center')
            c.border=bdr
            if ri % 2 == 0:
                c.fill=PatternFill('solid',fgColor='F8F9FA')

    for i, w in enumerate([8,12,16,10,24,1:,3:,10,18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────
# HTML 템플릿 (동문 제출 폼)
# ──────────────────────────────────────────
HTML_FORM = r"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>개인정보 수집 및 이용 동의서 | 덕성여자대학교</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/signature_pad@4.1.7/dist/signature_pad.umd.min.js"></script>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Noto Sans KR',sans-serif;background:linear-gradient(135deg,#1a1a2e 0%,#16213e 50%,#0f3460 100%);min-height:100vh;padding:40px 20px}
.container{max-width:720px;margin:0 auto;background:#fff;border-radius:20px;box-shadow:0 25px 80px rgba(0,0,0,.4);overflow:hidden}
.header{background:linear-gradient(135deg,#0f3460 0%,#e94560 100%);color:#fff;padding:40px;text-align:center}
.header .logo{font-size:13px;opacity:.75;letter-spacing:2px;text-transform:uppercase;margin-bottom:12px}
.header h1{font-size:22px;font-weight:700;letter-spacing:-0.5px;margin-bottom:8px}
.header p{font-size:13px;opacity:.8;line-height:1.7}
.notice{background:#f0f4ff;border-left:4px solid #0f3460;margin:28px 40px;padding:18px 20px;border-radius:0 10px 10px 0;font-size:13px;color:#444;line-height:1.9}
.notice-title{font-weight:700;color:#0f3460;margin-bottom:10px;font-size:14px}
.notice ol{padding-left:20px}
.notice li{margin-bottom:3px}
.form-body{padding:0 40px 40px}
.section{margin-bottom:28px}
.section-title{font-size:14px;font-weight:700;color:#1a1a2e;margin-bottom:16px;padding-bottom:8px;border-bottom:2px solid #ecf0f1;display:flex;align-items:center;gap:8px}
.section-title::before{content:'';width:4px;height:18px;background:linear-gradient(#0f3460,#e94560);border-radius:2px;display:inline-block}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.grid.full{grid-template-columns:1fr}
.field{display:flex;flex-direction:column;gap:6px}
label{font-size:12px;font-weight:600;color:#555}
label .req{color:#e94560;margin-left:2px}
input{padding:12px 14px;border:1.5px solid #dde2e8;border-radius:8px;font-size:14px;font-family:'Noto Sans KR',sans-serif;color:#333;background:#fafafa;transition:all .2s}
input:focus{outline:none;border-color:#0f3460;box-shadow:0 0 0 3px rgba(15,52,96,.12);background:#fff}
.consent-row{display:flex;gap:14px;margin-top:4px}
.consent-label{flex:1;display:flex;align-items:center;justify-content:center;gap:8px;padding:12px;border:2px solid #dde2e8;border-radius:10px;cursor:pointer;font-weight:600;font-size:14px;transition:all .2s}
.consent-label input[type=radio]{display:none}
.consent-label:has(input:checked){border-color:#27ae60;background:#f0fff8;color:#27ae60}
.consent-label.no:has(input:checked){border-color:#e74c3c;background:#fff5f5;color:#e74c3c}
.sig-box{border:1.5px solid #dde2e8;border-radius:10px;overflow:hidden}
.sig-bar{display:flex;justify-content:space-between;align-items:center;padding:8px 14px;background:#f7f8fa;border-bottom:1px solid #e8e8e8;font-size:12px;color:#999}
#sigCanvas{display:block;width:100%;height:150px;cursor:crosshair;touch-action:none;background:#fafafa}
.btn-clear{background:none;border:1px solid #ccc;border-radius:5px;padding:4px 10px;font-size:11px;cursor:pointer;color:#666;font-family:'Noto Sans KR',sans-serif}
.submit-btn{width:100%;padding:17px;background:linear-gradient(135deg,#0f3460,#e94560);color:#fff;border:none;border-radius:12px;font-size:16px;font-weight:700;font-family:'Noto Sans KR',sans-serif;cursor:pointer;letter-spacing:.5px;transition:transform .15s,box-shadow .15s;margin-top:6px}
.submit-btn:hover{transform:translateY(-2px);box-shadow:0 10px 25px rgba(233,69,96,.35)}
.submit-btn:disabled{opacity:.7;transform:none}
.success-banner{display:none;background:#f0fff8;border:2px solid #27ae60;border-radius:12px;padding:18px 20px;margin-top:20px;text-align:center;color:#27ae60;font-weight:600;font-size:15px}
@media(max-width:600px){.header,.form-body{padding-left:20px;padding-right:20px}.notice{margin:20px}.grid{grid-template-columns:1fr}}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <div class="logo">Duksung Women's University</div>
    <h1>개인정보 수집 및 이용 동의서</h1>
    <p>아래 양식을 작성하시면 동의서 PDF가 자동 발급됩니다.</p>
  </div>
  <div class="notice">
    <div class="notice-title">📋 개인정보 수집·이용 안내</div>
    <ol>
      <li><b>수집·이용자</b>: 덕성여자대학교</li>
      <li><b>수집·이용 목적</b>: 동문 관리, 학교 행사 및 소식 안내 등</li>
      <li><b>수집 항목(필수)</b>: 전공(학과), 입학연도(학번), 성명, 이메일, 핸드폰번호, 주소</li>
      <li><b>보유 기간</b>: 기록 보존이 필요한 시기까지</li>
    </ol>
    <p style="margin-top:8px;color:#e74c3c;font-size:12px">※ 동의 거부 시 동문 관리 및 행사·소식 안내 대상에서 제외될 수 있습니다.</p>
  </div>
  <div class="form-body">
    <form id="form" onsubmit="return onSubmit(event)">
      <div class="section">
        <div class="section-title">기본 정보</div>
        <div class="grid">
          <div class="field"><label>입학연도(학번)<span class="req">*</span></label>
            <input name="yaar" placeholder="예) 2001 또는 01학번" required></div>
          <div class="field"><label>성명<span class="req">*</span></label>
            <input name="name" placeholder="홍길동" required></div>
          <div class="field"><label>이메일<span class="req">*</span></label>
            <input name="email" type="email" placeholder="example@email.com" required></div>
          <div class="field"><label>핸드폰<span class="req">*</span></label>
            <input name="phone" type="tel" placeholder="010-0000-0000" required></div>
        </div>
        <div class="grid full" style="margin-top:14px">
          <div class="field"><label>주소<span class="req">*</span></label>
            <input name="address" placeholder="서울시 도봉구 ..." required></div>
        </div>
      </div>
      <div class="section">
        <div class="section-title">개인정보 수집·이용 동의</div>
        <div class="consent-row">
          <label class="consent-label"><input type="radio" name="consent" value="yes" required> ✅ 동의합니다</label>
          <label class="consent-label n"><input type="radio" name="consent" value="no"> ❌ 동의하지 않습니다</label>
        </div>
      </div>
      <div class="section">
        <div class="section-title">서명</div>
        <div class="sig-box">
          <div class="sig-bar"><span>아래 공간에 서명해 주세요</span>
            <button type="button" class="btn-clear" onclick="pad.clear()">지우기</button></div>
          <canvas id="sigCanvas"></canvas>
        </div>
        <input type="hidden" name="signature" id="sigData">
      </div>
      <button type="submit" class="submit-btn" id="btn">✉️ 동의서 제출 및 PDF 다운로드</button>
      <div class="success-banner" id="banner">✅ 제출이 완료되었습니다! PDF가 다운로드 중입니다.</div>
    </form>
  </div>
</div>
<script>
const canvas=document.getElementById('sigCanvas');
const pad=new SignaturePad(canvas,{backgroundColor:'rgb(250,250,250)',penColor:'rgb(20,20,80)',minWidth:1.5,maxWidth:3});
function resize(){const r=Math.max(window.devicePixelRatio||1,1);canvas.width=canvas.offsetWidth*r;canvas.height=150*r;canvas.getContext('2d').scale(r,r);pad.clear();}
window.addEventListener('resize',resize);resize();
async function onSubmit(e){
  e.preventDefault();
  if(!pad.isEmpty())document.getElementById('sigData').value=pad.toDataURL('image/png');
  const fd=new FormData(e.target),obj={};fd.forEach((v,k)=>obj[k]=v);
  const btn=document.getElementById('btn');btn.textContent='처리 중...';btn.disabled=true;
  try{
    const res=await fetch('/submit',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(obj)});
    if(res.ok){
      const blob=await res.blob();
      const a=document.createElement('a');a.href=URL.createObjectURL(blob);
      a.download=`개인정보동의서_${obj.name||'동문'}.pdf`;a.click();
      btn.style.background='linear-gradient(135deg,#27ae60,#2ecc71)';
      btn.textContent='✅ 제출 완료!';
      document.getElementById('banner').style.display='block';
      e.target.reset();pad.clear();
    }else{alert('오류: '+(await res.text()));btn.textContent='✉️ 동의서 제출 및 PDF 다운로드';btn.disabled=false;}
  }catch(err){alert('오류: '+err);btn.textContent='✉️ 동의서 제출 및 PDF 다운로드';btn.disabled=false;}
}
</script>
</body>
</html>"""


# ──────────────────────────────────────────
# 관리자 페이지
# ──────────────────────────────────────────
def admin_html(rows):
    count  = len(rows)
    agreed = sum(1 for r in rows if r[6]=='동의')
    rows_html = ''.join(f"""<tr>
      <td>{r[0]}</td><td>{r[1]}</td><td>{r[2]}</td><td>{r[3]}</td>
      <td>{r[4]}</td><td>{r[5]}</td>
      <td class="{'agree' if r[6]=='동의' else 'disagree'}">{r[6]}</td><td>{r[7]}</td>
      <td>{'<a href="/admin/pdf/'+str(r[0])+'?pw={ADMIN_PASS}" class="pdf-btn">⬇ PDF</a>'.format(ADMIN_PASS=ADMIN_PASS) if r[8] else '<span style="color:#bbb">없음</span>'}</td>
    </tr>""" for r in rows)
    return f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>관리자 | 동문 명단</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;600;700&display=swap" rel="stylesheet">
<style>
body{{font-family:'Noto Sans KR',sans-serif;background:#f0f2f5;margin:0;padding:30px 20px}}
.wrap{{max-width:1200px;margin:0 auto}}
h1{{font-size:22px;color:#1a1a2e;margin-bottom:6px}}
.sub{{color:#888;font-size:13px;margin-bottom:24px}}
.cards{{display:flex;gap:16px;margin-bottom:24px;flex-wrap:wrap}}
.card{{background:#fff;border-radius:12px;padding:20px 28px;box-shadow:0 2px 8px rgba(0,0,0,.08);min-width:150px}}
.card .num{{font-size:32px;font-weight:700;color:#0f3460}}
.card .lbl{{font-size:12px;color:#999;margin-top:4px}}
.dl-btn{{background:linear-gradient(135deg,#0f3460,#e94560);color:#fff;border:none;padding:12px 28px;border-radius:8px;font-size:14px;font-weight:700;cursor:pointer;font-family:'Noto Sans KR',sans-serif;text-decoration:none;display:inline-block;margin-bottom:20px}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08)}}
th{{background:#1a1a2e;color:#fff;padding:12px 14px;font-size:12px;text-align:center}}
td{{padding:10px 14px;font-size:13px;border-bottom:1px solid #f0f0f0;text-align:center}}
tr:last-child td{{border-bottom:none}}
tr:hover td{{background:#f8f9ff}}
.agree{{color:#27ae60;font-weight:600}}.disagree{{color:#e74c3c;font-weight:600}}
.pdf-btn{{background:#0f3460;color:#fff;padding:4px 10px;border-radius:6px;text-decoration:none;font-size:12px;font-weight:600}}
.pdf-btn:hover{{background:#e94560}}
</style></head>
<body><div class="wrap">
  <h1>동문 개인정보 동의 현황</h1>
  <div class="sub">관리자 전용 페이지</div>
  <div class="cards">
    <div class="card"><div class="num">{count}</div><div class="lbl">총 제출 수</div></div>
    <div class="card"><div class="num" style="color:#27ae60">{agreed}</div><div class="lbl">동의</div></div>
    <div class="card"><div class="num" style="color:#e74c3c">{count-agreed}</div><div class="lbl">미동의</div></div>
  </div>
  <a href="/admin/export?pw={ADMIN_PASS}" class="dl-btn">⬇ Excel 다운로드</a>
  <table><thead><tr>
    <th>No</th><th>입학연도</th><th>성명</th><th>이메일</th>
    <th>핸드폰</th><th>주소</th><th>동의여부</th><th>제출일시</th><th>PDF</th>
  </tr></thead>
  <tbody>{rows_html if rows_html else '<tr><td colspan="9" style="color:#aaa;padding:30px">아직 제출된 데이터가 없습니다.</td></tr>'}</tbody>
  </table>
</div></body></html>"""


# ──────────────────────────────────────────
# Tornado 핸들러
# ──────────────────────────────────────────
class MainHandler(tornado.web.RequestHandler):
    def get(self):
        self.set_header('Content-Type','text/html; charset=utf-8')
        self.write(HTML_FORM)

class SubmitHandler(tornado.web.RequestHandler):
    def post(self):
        try:
            data = json.loads(self.request.body)
        except Exception:
            self.set_status(400); self.write('잘못된 요청'); return
        try:
            pdf_bytes = make_consent_pdf(data)
        except Exception as e:
            self.set_status(500); self.write(f'PDF 오류: {e}'); return
        try:
            save_submission(data, pdf_bytes)
        except Exception as e:
            print(f'[DB 오류] {e}')
        try:
            send_notification_email(data, pdf_bytes)
        except Exception as e:
            print(f'[이메일 오류] {e}')
        fn = tornado.escape.url_escape(f"개인정보동의서_{data.get('name','동문')}.pdf")
        self.set_header('Content-Type','application/pdf')
        self.set_header('Content-Disposition',f"attachment; filename*=UTF-8''{fn}")
        self.write(pdf_bytes)

class AdminHandler(tornado.web.RequestHandler):
    def get(self):
        if self.get_argument('pw','') != ADMIN_PASS:
            self.set_header('Content-Type','text/html; charset=utf-8')
            self.write("""<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR&display=swap" rel="stylesheet">
<style>body{font-family:'Noto Sans KR',sans-serif;display:flex;justify-content:center;align-items:center;min-height:100vh;background:#f0f2f5;margin:0}
.box{background:#fff;padding:40px;border-radius:16px;box-shadow:0 4px 20px rgba(0,0,0,.1);text-align:center;width:320px}
h2{margin-bottom:20px;color:#1a1a2e;font-size:18px}
input{width:100%;padding:12px;border:1.5px solid #dde2e8;border-radius:8px;font-size:14px;font-family:'Noto Sans KR',sans-serif;margin-bottom:12px}
button{width:100%;padding:12px;background:#0f3460;color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:700;cursor:pointer;font-family:'Noto Sans KR',sans-serif}
</style></head><body>
<div class="box"><h2>🔒 관리자 로그인</h2>
<form method="get"><input type="password" name="pw" placeholder="비밀번호 입력">
<button type="submit">로그인</button></form></div></body></html>""")
            return
        self.set_header('Content-Type','text/html; charset=utf-8')
        self.write(admin_html(get_all_submissions()))

class AdminExportHandler(tornado.web.RequestHandler):
    def get(self):
        if self.get_argument('pw','') != ADMIN_PASS:
            self.set_status(403); self.write('인증 필요'); return
        today = datetime.date.today().strftime('%Y%m%d')
        fn = tornado.escape.url_escape(f'동문명단_{today}.xlsx')
        self.set_huader('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.set_header('Content-Disposition',f"attachment; filename*=UTF-8''{fn}")
        self.write(export_excel())

class AdminPDFHandler(tornado.web.RequestHandler):
    def get(self, sub_id):
        if self.get_argument('pw','') != ADMIN_PASS:
            self.set_status(403); self.write('인증 필요'); return
        con = get_conn()
        cur = con.cursor()
        cur.execute("SELECT name, pdf_data FROM submissions WHERE id=%s", (sub_id,))
        row = cur.fetchone()
        cur.close()
        con.close()
        if not row or not row[1]:
            self.set_status(404); self.write('PDF 없음'); return
        name, pdf_data = row
        fn = tornado.escape.url_escape(f"개인정보동의서_{name}.pdf")
        self.set_header('Content-Type','application/pdf')
        self.set_header('Content-Disposition',f"attachment; filename*=UTF-8''{fn}")
        self.write(bytes(pdf_data))

def make_app():
    return tornado.web.Application([
        (r'/',                  MainHandler),
        (r'/submit',            SubmitHandler),
        (r'/admin',             AdminHandler),
        (r'/admin/export',      AdminExportHandler),
        (r'/admin/pdf/(\d+)',   AdminPDFHandler),
    ])

if __name__ == '__main__':
    init_db()
    PORT = int(os.environ.get('PORT', 8080))
    app  = make_app()
    app.listen(PORT, address='0.0.0.0')
    print(f'✅ 서버 시작: http://0.0.0.0:{PORT}')
    print(f'   관리자: http://localhost:{PORT}/admin?pw={ADMIN_PASS}')
    tornado.ioloop.IOLoop.current().start()
